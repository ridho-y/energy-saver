from calendar import month_name
import openpyxl
from datetime import datetime, timedelta

############ FUNCTIONS ##############

def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days + 1)):
        yield start_date + timedelta(n)

def dayCost(date, hour):
    total = daily_charge

############ FUNCTIONS ##############

prices = {'summer': {}, 'non_summer': {}, 'winter': {}, 'non-winter': {}}

COLUMNS = 48

# DATE START & DATE END
date_start = datetime(2022, 1, 1)
date_end = datetime(2022, 5, 21)
ROWS = (date_end - date_start).days + 1

# DAILY CHARGE
daily_charge = 1.76
# daily_charge = input('Daily charge ($): ')

# FEED IN TARIFF
solar_price = 0.10
# solar_price = input('Solar feed in tariff ($): ')

# SOLAR KWH
# solar_kwh = input('Solar energy collected (kWh) during this period: ')

# Input Prices
print('Weekday summer prices (c/kWh)')
p = [int(i) for i in input('{peak} {tS} {tE} {off-peak} {tS} {tE} {shoulder} {tS} {tE}: ').split(' ')]
assert (len(p) == 9), 'Not all values given'
prices['summer']['weekday'] = {'peak': {'price': p[0], 'time_start': p[1], 'time_end': p[2]}, 'off-peak': {'price': p[3], 'time_start': p[4], 'time_end': p[5]}, 'shoulder': {'price': p[6], 'time_start': p[7], 'time_end': p[8]}}

print('Weekend summer prices (c/kWh)')
summer_weekend = input('{off-peak} {tS} {tE} {shoulder} {tS} {tE}: ').split(' ')
assert (len(p) == 4), 'Not all values given'
prices['summer']['weekend'] = {'off-peak': {'price': p[0], 'time_start': p[1], 'time_end': p[2]}, 'shoulder': {'price': p[3], 'time_start': p[4], 'time_end': p[5]}}

# print('Please input the non summer (1 April to 31 May) prices (c/kWh) in the following format:')
# non_summer = input('{off-peak} {shoulder}').split(' ')

# print('Please input the weekday winter prices (c/kWh) in the following format:')
# winter_weekday = input('{peak} {off-peak} {shoulder}').split(' ')

# print('Please input the weekend winter prices (c/kWh) in the following format:')
# winter_weekend = input('{off-peak} {shoulder}').split(' ')

# print('Please input the non winter (1 September to 31 October) prices (c/kWh) in the following format:')
# non_winter = input('{off-peak} {shoulder}').split(' ')


##################################################################################################################################################################

# summer = ['nov', 'dec', 'jan', 'feb', 'mar']
# s_non_seasonal = ['apr', 'may']
# winter = ['jun', 'jul', 'aug']
# w_non_seasonal = ['sep', 'oct']

# if month in summer:
#     season = 'summer'
# elif month in s_non_seasonal:
#     season = 'autumn'
# elif month in winter:
#     season = 'winter'
# else:
#     season = 'spring'


# ##########################

# path = "energy use.xlsx"
# wb_obj = openpyxl.load_workbook(path)
# sheet_obj = wb_obj.active

# assert (sheet_obj.max_column == COLUMNS), f'Incorrect number of columns: {sheet_obj.max_column} (expected: {COLUMNS}), cannot calculate'
# assert (sheet_obj.max_row == ROWS), f'Incorrect number of rows: {sheet_obj.max_row} (expected: {ROWS}), does not match date difference, cannot calculate'

# row_day = 1
# for single_date in daterange(date_start, date_end):
#     print(single_date.strftime("%Y-%m-%d"))
    
#     for t in range(0, COLUMNS):
#         time = t + 1
#         print(sheet_obj.cell(row=row_day,column=time).value, end=" ")
#         print(t // 2)


#     print('--')
#     print('--')
#     row_day += 1

# # total_costings = 



# # cell_obj = sheet_obj.cell(row = 3, column = 1)
# # print(type(cell_obj.value))
# # print(cell_obj.value)

